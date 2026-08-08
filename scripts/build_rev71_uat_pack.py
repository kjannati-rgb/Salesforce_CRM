from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

F = "Arial"
NAVY = "1F3864"; BLUE = "2E5496"; LT = "D9E1F2"; GREY = "F2F2F2"
GREEN = "C6EFCE"; RED = "FFC7CE"; AMBER = "FFEB9C"; WHITE = "FFFFFF"
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def cell(ws, ref, val, *, bold=False, size=10, color="000000", fill=None,
         wrap=False, align="left", valign="top", border_on=True):
    c = ws[ref]; c.value = val
    c.font = Font(name=F, bold=bold, size=size, color=color)
    c.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
    if fill: c.fill = PatternFill("solid", fgColor=fill)
    if border_on: c.border = border
    return c

wb = Workbook()

# ============ TAB 1: OVERVIEW ============
ov = wb.active; ov.title = "UAT Overview"
ov.sheet_view.showGridLines = False
ov.column_dimensions["A"].width = 3
for col in "BCDE": ov.column_dimensions[col].width = 26

cell(ov, "B2", "REV-71  ·  ALM Invoice Consolidation Code", bold=True, size=16, color="FFFFFF", fill=NAVY, border_on=False)
ov.merge_cells("B2:E2"); ov.row_dimensions[2].height = 28
cell(ov, "B3", "User Acceptance Testing (UAT) Pack", bold=True, size=12, color="FFFFFF", fill=BLUE, border_on=False)
ov.merge_cells("B3:E3"); ov.row_dimensions[3].height = 20

rows = [
    ("Environment", "KJDEV sandbox"),
    ("Login URL", "https://lawbusinessresearch--kjdev.sandbox.my.salesforce.com"),
    ("Field to check", "ALM Total Contract Value  (on each product line)"),
    ("Prepared", "18 Jun 2026"),
    ("Owner / approver", "Cayla Vichot — ALM Sales Ops"),
]
r = 5
for k, v in rows:
    cell(ov, f"B{r}", k, bold=True, size=10, fill=GREY)
    cell(ov, f"C{r}", v, size=10); ov.merge_cells(f"C{r}:E{r}")
    r += 1

r += 1
cell(ov, f"B{r}", "What we are testing", bold=True, size=11, color="FFFFFF", fill=BLUE); ov.merge_cells(f"B{r}:E{r}"); r += 1
purpose = ("When a deal has more than one product, the system should automatically put a 1, 2 or 3 "
           "in the ALM Total Contract Value field on every line, so Integra rolls the deal onto one "
           "invoice line. This pack checks it picks the right code on every type of deal — and, just "
           "as importantly, that it leaves the field blank when it should.")
cell(ov, f"B{r}", purpose, size=10, wrap=True); ov.merge_cells(f"B{r}:E{r}"); ov.row_dimensions[r].height = 58; r += 2

cell(ov, f"B{r}", "How to run each test", bold=True, size=11, color="FFFFFF", fill=BLUE); ov.merge_cells(f"B{r}:E{r}"); r += 1
steps = [
    "1.  Open the 'Test Cases' tab and pick a row.",
    "2.  Build / edit the deal exactly as the 'How to set it up' column says.",
    "3.  Save, then look at the ALM Total Contract Value field on the product lines.",
    "4.  Compare it to 'Expected result'. Put what you actually saw in 'Actual result'.",
    "5.  Set 'Status' to Pass or Fail (dropdown). Add your name, the date, and any notes.",
]
for s in steps:
    cell(ov, f"B{r}", s, size=10, wrap=True); ov.merge_cells(f"B{r}:E{r}"); ov.row_dimensions[r].height = 16; r += 1
r += 1

cell(ov, f"B{r}", "Live progress", bold=True, size=11, color="FFFFFF", fill=BLUE); ov.merge_cells(f"B{r}:E{r}"); r += 1
G = "'Test Cases'!$G$2:$G$23"
prog = [("Total tests", '=COUNTA(\'Test Cases\'!$A$2:$A$23)', None),
        ("Passed", f'=COUNTIF({G},"Pass")', GREEN),
        ("Failed", f'=COUNTIF({G},"Fail")', RED),
        ("Blocked", f'=COUNTIF({G},"Blocked")', AMBER),
        ("Not run yet", f'=COUNTIF({G},"Not Run")+SUMPRODUCT(--({G}=""))', GREY)]
for label, formula, fill in prog:
    cell(ov, f"B{r}", label, bold=True, size=10, fill=fill if fill else WHITE)
    cell(ov, f"C{r}", formula, size=10, align="center"); ov.merge_cells(f"C{r}:E{r}")
    r += 1
r += 1

cell(ov, f"B{r}", "Sign-off", bold=True, size=11, color="FFFFFF", fill=BLUE); ov.merge_cells(f"B{r}:E{r}"); r += 1
for label in ["UAT completed by", "Outcome (Approved / Rejected)", "Date", "Comments"]:
    cell(ov, f"B{r}", label, bold=True, size=10, fill=GREY)
    cell(ov, f"C{r}", "", size=10); ov.merge_cells(f"C{r}:E{r}"); ov.row_dimensions[r].height = 20
    r += 1

# ============ TAB 2: TEST CASES ============
tc = wb.create_sheet("Test Cases")
tc.sheet_view.showGridLines = False
headers = ["ID", "Group", "Scenario (what you're testing)", "How to set it up",
           "Expected result", "Actual result", "Status", "Tester", "Date", "Comments"]
widths = [6, 24, 30, 46, 34, 26, 11, 14, 12, 28]
for i, (h, w) in enumerate(zip(headers, widths), start=1):
    c = tc.cell(row=1, column=i, value=h)
    c.font = Font(name=F, bold=True, size=10, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border
    tc.column_dimensions[chr(64+i)].width = w
tc.row_dimensions[1].height = 30

cases = [
 ("A1","A — Core codes","Law.com worth more than International","Create a deal with Law.com + Law.com International. Price Law.com higher.","Code 1 on every product line"),
 ("A2","A — Core codes","International worth more than Law.com","Same two products; price Law.com International higher.","Code 3 on every line"),
 ("A3","A — Core codes","Equal value (the tie)","Same two products, priced exactly equal.","Code 1 (Premium wins the tie — confirmed by Brady)"),
 ("A4","A — Core codes","Global Leaders is on the deal","Add Global Leaders (GLBM) alongside any other product.","Code 2 on every line (Global Leaders always wins)"),
 ("A5","A — Core codes","Law.com + a small add-on (no International)","Law.com plus a smaller add-on (e.g. NV). No International.","Code 1"),
 ("A6","A — Core codes","International + a small add-on (no Law.com)","Law.com International plus a smaller add-on. No Law.com.","Code 3"),
 ("B1","B — Correctly blank","A single product on its own","A deal with only ONE product (e.g. just Law.com International).","Blank — no code. CONFIRMED correct: one line has nothing to consolidate."),
 ("B2","B — Correctly blank","A single bundle on its own","One bundle product, nothing else.","Blank — no code"),
 ("B3","B — Correctly blank","Several products, none a core brand","Two products that are NOT Law.com, International or Global Leaders.","Blank — and a 'no anchor' note is logged for review"),
 ("C1","C — Keeps up live","Add a second product","Start from a single-product deal (blank), then add a second qualifying product.","Code now appears on the lines"),
 ("C2","C — Keeps up live","Remove down to one product","Take a coded multi-product deal and delete lines until one remains.","Code clears back to blank"),
 ("C3","C — Keeps up live","Re-price so the bigger brand flips","On a Law.com + International deal, change pricing so the other brand becomes larger.","Code flips automatically (1 to 3, or 3 to 1)"),
 ("D1","D — Every sales path","Built through a Quote (CPQ)","Build the deal on a Quote, add lines, save / calculate.","Code appears on the quote AND carries down to the opportunity"),
 ("D2","D — Every sales path","Products added straight to the Opportunity","Add products directly on an Opportunity (no quote).","Code appears"),
 ("D3","D — Every sales path","Renewal deal","A renewal-type deal with qualifying products.","Code applied per the rules"),
 ("D4","D — Every sales path","Amendment deal","An amendment-type deal with qualifying products.","Code applied per the rules"),
 ("E1","E — Volume & safety","Large deal / many lines","A deal with many product lines (and/or several quotes).","Every line carries the same, correct code"),
 ("E2","E — Volume & safety","Bundle with components","A bundle that has child / component lines under a main product.","Only the main lines are coded; child components stay blank"),
 ("E3","E — Volume & safety","Re-save with no changes","Open a correctly-coded deal and save again without changing anything.","Nothing changes — no re-stamp, no flicker"),
 ("E4","E — Volume & safety","(Admin, optional) Emergency off-switch","Admin turns the kill switch on, then edits a deal.","No codes are added or changed while the switch is on"),
]
gfill = {"A — Core codes": "DDEBF7", "B — Correctly blank": "FCE4D6",
         "C — Keeps up live": "E2EFDA", "D — Every sales path": "FFF2CC",
         "E — Volume & safety": "EDEDED"}
row = 2
for cid, grp, scen, setup, exp in cases:
    vals = [cid, grp, scen, setup, exp, "", "", "", "", ""]
    for i, v in enumerate(vals, start=1):
        c = tc.cell(row=row, column=i, value=v)
        c.font = Font(name=F, size=10, bold=(i==1))
        c.alignment = Alignment(horizontal=("center" if i in (1,7,9) else "left"),
                                vertical="top", wrap_text=(i in (3,4,5,6,10)))
        c.border = border
        if i == 2: c.fill = PatternFill("solid", fgColor=gfill[grp])
        if i in (6,7,8,9,10): c.fill = PatternFill("solid", fgColor=WHITE)
    tc.row_dimensions[row].height = 42
    row += 1

tc.freeze_panes = "A2"
last = row - 1
dv = DataValidation(type="list", formula1='"Pass,Fail,Blocked,Not Run"', allow_blank=True)
tc.add_data_validation(dv); dv.add(f"G2:G{last}")
tc.conditional_formatting.add(f"G2:G{last}", CellIsRule(operator="equal", formula=['"Pass"'], fill=PatternFill("solid", fgColor=GREEN)))
tc.conditional_formatting.add(f"G2:G{last}", CellIsRule(operator="equal", formula=['"Fail"'], fill=PatternFill("solid", fgColor=RED)))
tc.conditional_formatting.add(f"G2:G{last}", CellIsRule(operator="equal", formula=['"Blocked"'], fill=PatternFill("solid", fgColor=AMBER)))

# ============ TAB 3: RULES REFERENCE ============
rf = wb.create_sheet("Rules Reference")
rf.sheet_view.showGridLines = False
rf.column_dimensions["A"].width = 3
rf.column_dimensions["B"].width = 48
rf.column_dimensions["C"].width = 12
rf.column_dimensions["D"].width = 50
cell(rf, "B2", "The coding rules in plain English", bold=True, size=13, color="FFFFFF", fill=NAVY, border_on=False)
rf.merge_cells("B2:D2"); rf.row_dimensions[2].height = 24
for i, h in enumerate(["If the deal contains…", "Code", "Why"], start=2):
    c = rf.cell(row=4, column=i, value=h)
    c.font = Font(name=F, bold=True, size=10, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = border
rf.row_dimensions[4].height = 20
rules = [
 ("Global Leaders (GLBM) — anywhere on the deal","2","Global Leaders is always the primary product"),
 ("Law.com + International, Law.com worth more","1","Law.com Premium bundle"),
 ("Law.com + International, International worth more","3","Law.com International bundle"),
 ("Law.com + International, worth exactly the same","1","Premium wins the tie (confirmed by Brady, 15 Jun 2026)"),
 ("Law.com only (with or without add-ons, 2+ lines)","1","Law.com is the anchor"),
 ("International only (with or without add-ons, 2+ lines)","3","International is the anchor"),
 ("Several products but none is a core brand","(blank)","No anchor — left blank and logged for review"),
 ("A single product on its own","(blank)","Nothing to consolidate (confirmed correct, 18 Jun 2026)"),
]
rr = 5
for cond, code, why in rules:
    cell(rf, f"B{rr}", cond, size=10, wrap=True)
    cell(rf, f"C{rr}", code, size=10, bold=True, align="center")
    cell(rf, f"D{rr}", why, size=10, wrap=True)
    rf.row_dimensions[rr].height = 30; rr += 1
rr += 1
cell(rf, f"B{rr}", "Note: Law.com Pro and Mid-Market are not given a Total Contract Value code — they use different codes, and the system leaves them alone.", size=9, color="808080", wrap=True, border_on=False)
rf.merge_cells(f"B{rr}:D{rr}"); rf.row_dimensions[rr].height = 28

out = r"C:\sf-work\kjdev\REV71_UAT_Test_Pack.xlsx"
wb.save(out)
print("saved", out)
